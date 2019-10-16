# !/usr/bin/python3
import os
import sys
import warnings
from copy import copy
from difflib import SequenceMatcher
from tkinter import *
from tkinter import messagebox
from tkinter import scrolledtext
from tkinter import ttk
from tkinter.filedialog import askopenfilename
from PIL import Image, ImageTk, ImageOps
from openpyxl import Workbook
from openpyxl import load_workbook


def resource_path(relative_path):
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

#warnings.filterwarnings("ignore")


class gloVari():
    path1 = "No file selected\t"
    path2 = "No file selected\t"
    path3 = "No file selected\t"
    path4 = "No file selected\t"
    ab=[]
    try:
        wb2 = load_workbook(filename='TP MASTER.xlsx', data_only=True)
    except:
        msg = messagebox.showinfo('TP Master.xlsx Not Found', 'Please copy TP Master file and Re-Run the program.')
    wf = wb2['MBBSClearance']
    wb=Workbook()
    wb.create_sheet('Output')
    wb.remove(wb['Sheet'])
    ws=wb['Output']
    sheets = []
    matdict={}
    cat = {}
    length = 0
    sh = {}
    def dec(self):
        ab = ['A', 'B', 'C', 'D', 'E', 'F', 'G','H']
        self.wb1 = load_workbook(filename=self.path1, data_only=True)
        for i in ab:
            t = i + str(1)
            self.ws[t].value = self.wf[t].value
            self.ws[t].font = copy(self.wf[t].font)
            self.ws[t].border = copy(self.wf[t].border)
            self.ws[t].fill = copy(self.wf[t].fill)
            self.ws[t].number_format = copy(self.wf[t].number_format)
            self.ws[t].protection = copy(self.wf[t].protection)
            self.ws[t].alignment = copy(self.wf[t].alignment)
        self.ws1=self.wb1['Sheet1']
        l=self.ws1.max_row+1
        for i in range(2,l):
            c='C'+str(i)
            d='D'+str(i)
            q='E'+str(i)
            mat=self.ws1[c].value
            matd=self.ws1[d].value
            matq=self.ws1[q].value
            if mat in self.matdict:
                self.matdict[mat][2]+=matq
            else:
                self.matdict[mat]=[mat,matd,matq,0,0,0,0]
        self.wb2 = load_workbook(filename=self.path2, data_only=True)
        self.ws2=self.wb2['Sheet1']
        l=self.ws2.max_row
        for i in range(2,l):
            c='C'+str(i)
            d='D'+str(i)
            e='E'+str(i)
            mat=self.ws2[c].value
            matq=self.ws2[e].value
            matd=self.ws2[d].value
            if mat in self.matdict:
                self.matdict[mat][3] += matq
                self.matdict[mat][4] = self.matdict[mat][2]-self.matdict[mat][3]
            else:
                self.matdict[mat]=[mat,matd,0,matq,0,0,0]
        self.wb3 = load_workbook(filename=self.path3, data_only=True)
        self.ws3=self.wb3['Sheet1']
        l=self.ws3.max_row
        for i in range(2,l):
            c='A'+str(i)
            d='B'+str(i)
            ec='C'+str(i)
            ee='E'+str(i)
            ef='F'+str(i)
            mat=self.ws3[c].value
            matq=self.ws3[ec].value+self.ws3[ee].value+self.ws3[ef].value
            matd=self.ws3[d].value
            if mat in self.matdict:
                self.matdict[mat][5] += matq
            else:
                self.matdict[mat]=[mat,matd,0,0,0,matq,0]
        self.wb4 = load_workbook(filename=self.path4, data_only=True)
        self.ws4=self.wb4['Sheet1']
        l=self.ws4.max_row
        for i in range(2,l):
            c='G'+str(i)
            d='H'+str(i)
            e='I'+str(i)
            mat=self.ws4[c].value
            matq=self.ws4[e].value
            matd=self.ws4[d].value
            if mat in self.matdict:
                self.matdict[mat][6] += matq
            else:
                self.matdict[mat]=[mat,matd,0,0,0,0,matq]
        x=2
        for i in self.matdict:
           self.ws.append(self.matdict[i])
        self.ws.column_dimensions['A'].width = self.wf.column_dimensions['A'].width
        self.ws.column_dimensions['B'].width = self.wf.column_dimensions['B'].width

global txtbox

def savetp():
    try:
        gloVari.wb2.save(filename='TP MASTER.xlsx')
    except:
        msg = messagebox.showinfo('Close TP MASTER.XLX FILE', 'Please close the file and press OK to continue.')
        savetp()


def run(i):
    progressBar['maximum'] = 100
    x = 100 / gloVari.length
    progressBar["value"] = i * x
    progressBar.update()


def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()


def resize_image(event):
    new_width = event.width
    new_height = event.height
    image = copy_of_image.resize((new_width, new_height))
    photo = ImageTk.PhotoImage(image)
    label.config(image=photo)
    label.image = photo  # avoid garbage collection


def helloCallBack1():
    global label1
    global file1
    label1.destroy()
    gloVari.path1 = askopenfilename()
    print(gloVari.path1.rsplit('/', 1)[-1])
    file1 = gloVari.path1.rsplit('/', 1)[-1]
    label1 = Label(workzone, text=file1, bg='#dadfe3', font=("Arial Rounded MT Bold", 8), foreground="#173d5c")
    label1.grid(row=1, column=2, sticky='ENWS', padx=(0, 5), pady=(5, 0))
    #txtbox.delete(1.0, END)
    print(gloVari.path1)

def helloCallBack2():
    global label2
    global file2
    label2.destroy()
    gloVari.path2 = askopenfilename()
    print(gloVari.path2.rsplit('/', 1)[-1])
    file2 = gloVari.path2.rsplit('/', 1)[-1]
    label2 = Label(workzone, text=file2, bg='#dadfe3', font=("Arial Rounded MT Bold", 8), foreground="#173d5c")
    label2.grid(row=2, column=2, sticky='ENWS', padx=(0, 5), pady=(5, 0))
    #txtbox.delete(2.0, END)
    print(gloVari.path2)

def helloCallBack3():
    global label3
    global file3
    label3.destroy()
    gloVari.path3 = askopenfilename()
    print(gloVari.path3.rsplit('/', 1)[-1])
    file3 = gloVari.path3.rsplit('/', 1)[-1]
    label3 = Label(workzone, text=file3, bg='#dadfe3', font=("Arial Rounded MT Bold", 8), foreground="#173d5c")
    label3.grid(row=3, column=2, sticky='ENWS', padx=(0, 5), pady=(5, 0))
    txtbox.delete(3.0, END)
    print(gloVari.path3)

def helloCallBack4():
    global label4
    global file4
    label4.destroy()
    gloVari.path4 = askopenfilename()
    print(gloVari.path4.rsplit('/', 1)[-1])
    file4 = gloVari.path4.rsplit('/', 1)[-1]
    label4 = Label(workzone, text=file4, bg='#dadfe3', font=("Arial Rounded MT Bold", 8), foreground="#173d5c")
    label4.grid(row=4, column=2, sticky='ENWS', padx=(0, 5), pady=(5, 0))
    txtbox.delete(4.0, END)
    print(gloVari.path4)


def process():
    if gloVari.path1 == 'No file selected\t':
        print('Please select a file.')
        msg = messagebox.showinfo('Please select a file.', gloVari.path1)
    else:
        gloVari.dec(gloVari)
        print(gloVari.path1.rsplit('/', 1)[-1])
        txtbox.insert(INSERT, '\n\n  Please wait loading....\n\n\t')
        i = 0
        print("Loading...")
        txtbox.insert(INSERT, "\n\n\t\tTotal  Success!\n\t")
        saveop()

def saveop():
    file = gloVari.path1.rsplit('/', 1)[-1]
    try:
        gloVari.wb.save(filename='Output_' + file)
    except:
        msg = messagebox.showinfo('Close ' + 'Output_' + file, 'Please close the file and press OK to continue.')
        saveop()




root = Tk()
root.state('zoomed')
root.configure(background="#dadfe3")
root.title("TP Break Up")

# ========================================== Image URL =================================================================

title = resource_path("img/titlembbs.png")
img2 = resource_path("img/side.jpg")
lin = resource_path("img/line.png")
chfi = resource_path("img/choosefile.png")
ee1 = resource_path("img/edit.png")
chek = resource_path("img/check.png")

# =========================================== Opening Image ============================================================

image1 = Image.open(title)
imgr = Image.open(img2)
lini = Image.open(lin)
chck = Image.open(chek)
chofi = Image.open(chfi)
edi = Image.open(ee1)

# =========================================== Resizing =================================================================

photo1 = ImageTk.PhotoImage(image1)
image1 = image1.resize((1400, 179))
copy_of_image1 = image1.copy()
photo1 = ImageTk.PhotoImage(image1)

imgri = ImageTk.PhotoImage(imgr)
imgr = imgr.resize((460, 345))
copy_of_imager = imgr.copy()
imgri = ImageTk.PhotoImage(imgr)

line = ImageTk.PhotoImage(lini)
lini = lini.resize((600, 15))
copy_of_imagel = lini.copy()
line = ImageTk.PhotoImage(lini)

checkk = ImageTk.PhotoImage(chck)
chck = chck.resize((166, 41))
copy_of_imagec = chck.copy()
checkk = ImageTk.PhotoImage(chck)

chosfi = ImageTk.PhotoImage(chofi)
chofi = chofi.resize((139, 37))
copy_of_imagech = chofi.copy()
chosfi = ImageTk.PhotoImage(chofi)

editt = ImageTk.PhotoImage(edi)
edi = edi.resize((40, 40))
copy_of_imagee = edi.copy()
editt = ImageTk.PhotoImage(edi)

# =============================================== Frames ===============================================================

workzone = Frame(root, border='5', bg='grey')
workzone.grid(row=1, column=0, sticky='ENW', padx=(80, 10), pady=(20, 0), rowspan=2)

progress = Frame(root, border='5', bg='#dadfe3')
progress.grid(row=6, column=0, sticky='NS', padx=(10, 80))

group1 = LabelFrame(root, text="Details", padx=5, pady=5, bg='#dadfe3')
group1.grid(row=6, column=1, padx=(0,80), pady=(0, 45), sticky='NSEW')

# ============================================ Logo and Side Image =====================================================

title = Label(root, bg='#dadfe3', image=photo1)
title.grid(row=0, column=0, columnspan=2, sticky=N, pady=20)

sideimage = Label(root, bg='#dadfe3', image=imgri)
sideimage.grid(row=1, column=1, sticky='ENWS', padx=(0, 100), pady=(0, 0), rowspan=5)

# ============================================ Buttons and Labels ======================================================

lb1t = Label(workzone, text='Select Files', bg='#a0a0a0',font=("Arial Rounded MT Bold", 15), foreground="white")
lb1t.grid(row=0, column=0, sticky='ENWS', padx=(0, 5))
lb2t = Label(workzone, text='Selected File', bg='#a0a0a0',font=("Arial Rounded MT Bold", 15), foreground="white")
lb2t.grid(row=0, column=1, sticky='ENWS', padx=(0, 5), columnspan=2)
lb21 = Label(workzone, text='Status', bg='#a0a0a0',font=("Arial Rounded MT Bold", 15), foreground="white")
lb21.grid(row=0, column=3, sticky='ENWS')

# =========================================== Required by Design =======================================================

lb11 = Label(workzone, text='Required By Design\n[CS11]', bg='#dadfe3', border='5',font=("Arial Rounded MT Bold", 12), foreground="#173d5c")
lb11.grid(row=1, column=0, sticky='ENWS', padx=(0, 5), pady=(5,0))
label1 = Label(workzone, text=gloVari.path1, bg='#dadfe3',font=("Arial Rounded MT Bold", 8), foreground="#173d5c")
label1.grid(row=1, column=2, sticky='ENWS', padx=(0, 5), pady=(5,0))
buttonrd = Button(workzone, image=chosfi, bg='#dadfe3', border='0', command=helloCallBack1)
buttonrd.grid(row=1, column=1, padx=(0, 5), pady=(5,0), sticky='NEWS')
lb1 = Label(workzone, text='', bg='#dadfe3',font=("Arial Rounded MT Bold", 12), foreground="green")
lb1.grid(row=1, column=3, sticky='ENWS', pady=(5,0))

# ============================================== Posted [CJI3] =========================================================

lb12 = Label(workzone, text='Posted\n[CJI3]', bg='#dadfe3', border='5',font=("Arial Rounded MT Bold", 12), foreground="#173d5c")
lb12.grid(row=2, column=0, sticky='ENWS', padx=(0, 5), pady=(5,0))
label2 = Label(workzone, text=gloVari.path2, bg='#dadfe3',font=("Arial Rounded MT Bold", 8), foreground="#173d5c")
label2.grid(row=2, column=2, sticky='ENWS', padx=(0, 5), pady=(5,0))
buttonpo = Button(workzone, image=chosfi, bg='#dadfe3', border='0', command=helloCallBack2)
buttonpo.grid(row=2, column=1, padx=(0, 5), pady=(5,0), sticky='NEWS')
lb2 = Label(workzone, text='', bg='#dadfe3',font=("Arial Rounded MT Bold", 12), foreground="green")
lb2.grid(row=2, column=3, sticky='ENWS', pady=(5,0))


# =============================================  Project Stock [MBBS]  =========================================================

lb13 = Label(workzone, text='Project Stock \n[MBBS]', bg='#dadfe3', border='5',font=("Arial Rounded MT Bold", 12), foreground="#173d5c")
lb13.grid(row=3, column=0, sticky='ENWS', padx=(0, 5), pady=(5,0))
label3 = Label(workzone, text=gloVari.path3, bg='#dadfe3',font=("Arial Rounded MT Bold", 8), foreground="#173d5c")
label3.grid(row=3, column=2, sticky='ENWS', padx=(0, 5), pady=(5,0))
buttonps = Button(workzone, image=chosfi, bg='#dadfe3', border='0', command=helloCallBack3)
buttonps.grid(row=3, column=1, padx=(0, 5), pady=(5,0), sticky='NEWS')
lb3 = Label(workzone, text='', bg='#dadfe3',font=("Arial Rounded MT Bold", 12), foreground="green")
lb3.grid(row=3, column=3, sticky='ENWS', pady=(5,0))


# ============================================= OPEN PO [ME2J] =========================================================

lb14 = Label(workzone, text='OPEN PO \n[ME2J]', bg='#dadfe3', border='5',font=("Arial Rounded MT Bold", 12), foreground="#173d5c")
lb14.grid(row=4, column=0, sticky='ENWS', padx=(0, 5), pady=(5,0))
label4 = Label(workzone, text=gloVari.path4, bg='#dadfe3',font=("Arial Rounded MT Bold", 8), foreground="#173d5c")
label4.grid(row=4, column=2, sticky='ENWS', padx=(0, 5), pady=(5,0))
buttonopo = Button(workzone, image=chosfi, bg='#dadfe3', border='0', command=helloCallBack4)
buttonopo.grid(row=4, column=1, padx=(0, 5), pady=(5,0), sticky='NEWS')
lb4 = Label(workzone, text='', bg='#dadfe3',font=("Arial Rounded MT Bold", 12), foreground="green")
lb4.grid(row=4, column=3, sticky='ENWS', pady=(5,0))

# ======================================================================================================================

buttonbu = Button(root, image=checkk, bg='#dadfe3', border='0', command=process)
buttonbu.grid(row=3, column=0, pady=(20, 20), sticky='NS')

lbl = Label(root, image=line, bg='#dadfe3')
lbl.grid(row=4, column=0, sticky='ENWS', pady=20)

lbp = Label(progress, text='Progress: ', bg='#dadfe3')
lbp.grid(row=0, column=0, sticky='ENWS', pady=0)
progressBar = ttk.Progressbar(progress, orient="horizontal", length=286, mode="determinate")
progressBar.grid(row=0, column=1,pady=5)


# Create the textbox
txtbox = scrolledtext.ScrolledText(group1, width=30, height=30)
txtbox.grid(row=0, column=0, padx=0, pady=(0, 0), sticky='NSEW')

# ============================================= Column Weights =========================================================

root.columnconfigure(0, weight=15)
root.columnconfigure(1, weight=2)
workzone.columnconfigure(0, weight=2)
workzone.columnconfigure(1, weight=1)
workzone.columnconfigure(2, weight=3)
workzone.columnconfigure(3, weight=1)
group1.rowconfigure(0, weight=1)
group1.columnconfigure(0, weight=1)
root.rowconfigure(6, weight=1)

# ======================================================================================================================

root.mainloop()