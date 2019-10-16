# !/usr/bin/python3
import os
import sys
import warnings
from subprocess import call
from copy import copy
from difflib import SequenceMatcher
from tkinter import *
from tkinter import messagebox
from tkinter import scrolledtext
from tkinter import ttk
from tkinter.filedialog import askopenfilename

from PIL import Image, ImageTk, ImageOps
from openpyxl import Workbook
from openpyxl import load_workbook


def resource_path(relative_path):
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


# warnings.filterwarnings("ignore")


class gloVar():
    path = "No file selected\t"
    ab = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
          'W', 'X', 'Y', 'Z', 'AA', 'AB']
    abc = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
    try:
        wb2 = load_workbook(filename='TP MASTER.xlsx', data_only=True)
    except:
        msg = messagebox.showinfo('TP Master.xlsx Not Found', 'Please copy TP Master file and Re-Run the program.')
    ws3 = wb2['COST-CAT']
    ws4 = wb2['WO COSTCODE']
    wf = wb2['Formula']
    cmi = wb2['ConsMicro']
    cma = wb2['ConsMacro']
    pgoh = 1.142
    cats = {}
    dictc = {}
    dictwc = {}
    dictpwc = {}
    sybr = []
    dictmicro = {}
    dictmacro = {}
    j = 2
    for col in ws3.iter_rows(min_row=2):
        tempx = 'G' + str(j)
        y = ws3[tempx].value
        if y != None:
            y = y.upper()
        cats[y] = 0
        j += 1
    j = 2
    for col in ws4.iter_rows(min_row=2):
        tempx = 'B' + str(j)
        y = ws4[tempx].value
        if y != None:
            y = y.upper()
        cats[y] = 0
        j += 1
    wb1 = Workbook()
    wb = Workbook()
    ws2 = wb.active
    wb3 = Workbook()
    sheets = []
    cat = {}
    length = 0
    sh = {}

    def dec(self):
        self.wb = load_workbook(filename=self.path, data_only=True)
        try:
            self.wb1.remove(self.wb1['Sheet'])
        except:
            pass
        try:
            self.wb3.remove(self.wb3['Sheet'])
        except:
            pass
        for wx in self.wb.worksheets:
            wy = wx.title
            if wy.find('-FLI') != -1:
                self.ws2 = self.wb[wy]
            elif wy.find('SY') != -1:
                if len(wy) < 13:
                    self.wb1.create_sheet(wy)
                    self.wb3.create_sheet(wy + 'Mi')
                    self.wb3.create_sheet(wy + 'Ma')
                    self.sheets.append(wy)
                    self.length += 1
        for wx in self.wb1.worksheets:
            for i in self.ab:
                t = i + str(1)
                wx[t].value = self.wf[t].value
                wx[t].font = copy(self.wf[t].font)
                wx[t].border = copy(self.wf[t].border)
                wx[t].fill = copy(self.wf[t].fill)
                wx[t].number_format = copy(self.wf[t].number_format)
                wx[t].protection = copy(self.wf[t].protection)
                wx[t].alignment = copy(self.wf[t].alignment)
                wx.column_dimensions[i].width = self.wf.column_dimensions[i].width
        for wx in self.wb3.worksheets:
            for i in self.abc:
                t = i + str(1)
                u = i + str(2)
                wx[t].value = self.cmi[t].value
                wx[t].font = copy(self.cmi[t].font)
                wx[t].border = copy(self.cmi[t].border)
                wx[t].fill = copy(self.cmi[t].fill)
                wx[t].number_format = copy(self.cmi[t].number_format)
                wx[t].protection = copy(self.cmi[t].protection)
                wx[t].alignment = copy(self.cmi[t].alignment)
                wx.column_dimensions[i].width = self.cmi.column_dimensions[i].width

                tem = wx.title
                if tem.find('Mi') != -1:
                    wx[u].value = self.cmi[u].value
                elif tem.find('Ma') != -1:
                    wx[u].value = self.cma[u].value
                    print(self.cma[u].value)
                wx[u].font = copy(self.cmi[u].font)
                wx[u].border = copy(self.cmi[u].border)
                wx[u].fill = copy(self.cmi[u].fill)
                wx[u].number_format = copy(self.cmi[u].number_format)
                wx[u].protection = copy(self.cmi[u].protection)
                wx[u].alignment = copy(self.cmi[u].alignment)
                wx.column_dimensions[i].width = self.cmi.column_dimensions[i].width


global txtbox


class main(Frame):
    y = 0
    children_dict = dict()
    z = 0

    def newWindowOpen(self):
        childWindow = Toplevel()
        childWindow.rowconfigure(0, weight=1)
        childWindow.columnconfigure(0, weight=1)

        popCanv = Canvas(childWindow)  # width=1256, height = 1674)
        popCanv.grid(row=0, column=0, sticky="nsew")  # added sticky
        scroll = Scrollbar(childWindow, orient=VERTICAL)
        scroll.grid(row=0, column=1, sticky="ns")

        popCanv.rowconfigure(0, weight=1)
        popCanv.columnconfigure(0, weight=1)

        scroll.config(command=popCanv.yview)
        popCanv.configure(width=950)

        frame = Frame(popCanv, width=950)
        frame.grid(row=0, column=0, sticky="ns")

        popCanv.create_window((0, 0), window=frame, anchor='n', tags="my_tag")
        childWindow.wm_title("Enter Categories")
        popCanv.configure(yscrollcommand=scroll.set)

        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        itemLabel0 = Label(frame, text="Cost Code", font='Helvetica 10 bold')
        itemLabel1 = Label(frame, text="Item Description", font='Helvetica 10 bold')
        itemLabel2 = Label(frame, text="Category", font='Helvetica 10 bold')
        itemLabel3 = Label(frame, text="SS", font='Helvetica 10 bold')
        itemLabel4 = Label(frame, text="Al", font='Helvetica 10 bold')
        itemLabel5 = Label(frame, text="Cu", font='Helvetica 10 bold')
        itemLabel6 = Label(frame, text="HRS", font='Helvetica 10 bold')
        itemLabel0.grid(row=0, column=0)
        itemLabel1.grid(row=0, column=1)
        itemLabel2.grid(row=0, column=2)
        itemLabel3.grid(row=0, column=3)
        itemLabel4.grid(row=0, column=4)
        itemLabel5.grid(row=0, column=5)
        x = 0
        for i in gloVar.cat:
            j = StringVar()
            k = StringVar()
            l = StringVar()
            m = StringVar()
            itemLabel3 = Label(frame, text=gloVar.cat[i])
            itemLabel4 = Label(frame, text=i)
            itemEntry0 = Entry(frame, textvariable=j)
            itemEntry1 = Entry(frame, textvariable=k)
            itemEntry2 = Entry(frame, textvariable=l)
            itemEntry3 = Entry(frame, textvariable=m)

            self.children_dict[i] = [j, k, l, m]

            itemLabel3.grid(row=x + 1, column=0)
            itemLabel4.grid(row=x + 1, column=1)
            itemEntry0.grid(row=x + 1, column=2)
            itemEntry1.grid(row=x + 1, column=3)
            itemEntry2.grid(row=x + 1, column=4)
            itemEntry3.grid(row=x + 1, column=5)
            x += 1
            self.y = x
        for i in gloVar.dictpwc:
            j = StringVar()
            k = StringVar()
            l = StringVar()
            m = StringVar()
            itemLabel3 = Label(frame, text=gloVar.dictpwc[i])
            itemLabel4 = Label(frame, text=i)
            itemEntry0 = Entry(frame, textvariable=j)
            itemEntry1 = Entry(frame, textvariable=k)
            itemEntry2 = Entry(frame, textvariable=l)
            itemEntry3 = Entry(frame, textvariable=m)

            self.children_dict[i] = [j, k, l, m]

            itemLabel3.grid(row=x + 1, column=0)
            itemLabel4.grid(row=x + 1, column=1)
            itemEntry0.grid(row=x + 1, column=2)
            itemEntry1.grid(row=x + 1, column=3)
            itemEntry2.grid(row=x + 1, column=4)
            itemEntry3.grid(row=x + 1, column=5)
            x += 1
        self.z = x
        childWindow.submitButton = Button(frame, text="submit", command=lambda: self.submitTest(childWindow))
        childWindow.submitButton.grid(row=x + 2, pady=(10, 0), column=0, columnspan=5)
        x += 1
        # popCanv.configure(height=x * 15)
        frame.configure(height=x * 22)
        buff = 50 + x * 20
        childWindow.geometry("1000x" + str(buff))
        popCanv.configure(scrollregion=popCanv.bbox("all"))

    def submitTest(self, childWindow):
        for i in gloVar.cat:
            if gloVar.cat[i] is None or gloVar.cat[i] == 'N':
                gloVar.ws4.append(
                    [i, self.children_dict[i][0].get(), self.children_dict[i][1].get(), self.children_dict[i][2].get(),
                     self.children_dict[i][3].get()])
            else:
                gloVar.ws3.append([gloVar.cat[i], i, self.children_dict[i][1].get(), self.children_dict[i][2].get(),
                                   self.children_dict[i][3].get(), '', self.children_dict[i][0].get()])
        lws4 = gloVar.ws4.max_row + 1
        lws3 = gloVar.ws3.max_row + 1
        for i in gloVar.dictpwc:
            if gloVar.dictpwc[i] == 'N':
                for j in range(2, lws4):
                    loc = 'A' + str(j)
                    if gloVar.ws4[loc].value == i:
                        locc = 'C' + str(j)
                        locd = 'D' + str(j)
                        loce = 'E' + str(j)
                        try:
                            gloVar.ws4[locc] = self.children_dict[self.y][1].get()
                        except:
                            gloVar.ws4[locc] = None
                        try:
                            gloVar.ws4[locd] = self.children_dict[self.y][2].get()
                        except:
                            gloVar.ws4[locd] = None
                        try:
                            gloVar.ws4[loce] = self.children_dict[self.y][3].get()
                        except:
                            gloVar.ws4[loce] = None
                        self.y += 1
                        break
            else:
                for j in range(2, lws3):
                    loc = 'A' + str(j)
                    if gloVar.ws3[loc].value == gloVar.dictpwc[i]:
                        locc = 'C' + str(j)
                        locd = 'D' + str(j)
                        loce = 'E' + str(j)
                        try:
                            gloVar.ws3[locc] = self.children_dict[self.y][1].get()
                        except:
                            gloVar.ws3[locc] = None
                        try:
                            gloVar.ws3[locd] = self.children_dict[self.y][2].get()
                        except:
                            gloVar.ws3[locd] = None
                        try:
                            gloVar.ws3[loce] = self.children_dict[self.y][3].get()
                        except:
                            gloVar.ws3[loce] = None
                        self.y += 1
                        break
        savetp()
        childWindow.destroy()
        for i in gloVar.sh:
            bup(i)
        saveop()
        buttonbu = Button(root, image=conso, bg='#dadfe3', border='0', command=consolidate)
        buttonbu.grid(row=3, column=0, pady=(20, 20), sticky='NS')
        txtbox.insert(INSERT, "\n\n\t\tTotal  Success!\n\t")


def consolidate():
    for i in gloVar.sheets:
        shoMi = gloVar.wb3[i + 'Mi']
        shoMa = gloVar.wb3[i + 'Ma']
        sh = gloVar.wb1[i]
        l = sh.max_row + 1
        for j in range(2, l):
            a = 'A' + str(j)
            b = 'B' + str(j)
            e = 'E' + str(j)
            f = 'F' + str(j)
            jk = 'J' + str(j)
            catmicro = sh[b].value
            catmacro = sh[a].value
            if catmacro is None:
                catmacro = 0
            if catmicro is None:
                catmicro = 0
            qty = sh[e].value * sh[jk].value
            val = sh[f].value * qty / 3
            if catmicro in gloVar.dictmicro:
                gloVar.dictmicro[catmicro][0] += qty
                gloVar.dictmicro[catmicro][1] += val
            else:
                gloVar.dictmicro[catmicro] = [qty, val]
            if catmacro in gloVar.dictmacro:
                gloVar.dictmacro[catmacro][0] += qty
                gloVar.dictmacro[catmacro][1] += val
            else:
                gloVar.dictmacro[catmacro] = [qty, val]
        for i in gloVar.dictmicro:
            shoMi.append([i, gloVar.dictmicro[i][0], gloVar.dictmicro[i][1]])
        for i in gloVar.dictmacro:
            shoMa.append([i, gloVar.dictmacro[i][0], gloVar.dictmacro[i][1]])
        gloVar.dictmicro = {}
        gloVar.dictmacro = {}
    saveopc()
    txtbox.insert(INSERT, "\n\n\tAll process Complete!!\n\t")
    call(["excel", 'ConsolidatedOutput_' + gloVar.path.rsplit('/', 1)[-1]])
    sys.exit()


def saveopc():
    file = gloVar.path.rsplit('/', 1)[-1]
    try:
        gloVar.wb3.save(filename='ConsolidatedOutput_' + file)
    except:
        msg = messagebox.showinfo('Close ' + 'ConsolidatedOutput_' + file,
                                  'Please close the file and press OK to continue.')
        saveopc()

def savetp():
    try:
        gloVar.wb2.save(filename='TP MASTER.xlsx')
    except:
        msg = messagebox.showinfo('Close TP MASTER.XLX FILE', 'Please close the file and press OK to continue.')
        savetp()


def run(i):
    progressBar['maximum'] = 100
    x = 100 / gloVar.length
    progressBar["value"] = i * x
    progressBar.update()


def chk(a):
    if a is None:
        return 3
    elif a.find('Panel Code :') != -1:
        return 2
    elif a.find('Board Code : ') != -1:
        return 1
    else:
        return 0


def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()


def bup(wy):
    ws1 = gloVar.wb1[wy]
    l = ws1.max_row
    lws4 = gloVar.ws4.max_row
    lws3 = gloVar.ws3.max_row
    cc = ['H', 'C', 'D', 'E', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']
    wo = ['C', 'D', 'E', 'F', 'B', 'C', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']
    op = ['A', 'N', 'O', 'P', 'B', 'A', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB']
    for x in range(2, l + 1):
        t = 'I' + str(x)
        temp = ws1[t].value
        fy = 0
        ft = 0
        if temp is None or temp == 'N':
            ft = 1
            td = 'D' + str(x)
            tempd = ws1[td].value
            tempd = tempd.upper()
            for i in range(2, lws4 + 1):
                td1 = 'A' + str(i)
                tempd1 = gloVar.ws4[td1].value
                tempd1 = tempd1.upper()
                if tempd == tempd1:
                    fy = 1
                    td2 = 'B' + str(i)
                    td3 = 'B' + str(x)
                    c = 'C' + str(i)
                    d = 'D' + str(i)
                    e = 'E' + str(i)
                    v = 'N' + str(x)
                    w = 'O' + str(x)
                    z = 'P' + str(x)
                    for j, k in zip(wo, op):
                        ty = j + str(i)
                        tz = k + str(x)
                        ws1[tz] = gloVar.ws4[ty].value
                    xc = gloVar.ws4[c].value
                    xd = gloVar.ws4[d].value
                    xe = gloVar.ws4[e].value
                    catt = gloVar.ws4[td2].value
                    if catt is not None:
                        catt = catt.upper()
                    ws1[td3] = catt
                    if xc is None:
                        xc = 0
                    if xd is None:
                        xd = 0
                    if xe is None:
                        xe = 0
                    ws1[v].value = xc
                    ws1[w].value = xd
                    ws1[z].value = xe
                    if catt == 'LINK' or catt == 'MS':
                        if xc == 0 and xd == 0 and xe == 0:
                            gloVar.dictpwc[tempd] = 'N'
                            gloVar.sh[wy] = 0
                    break
        if ft == 0:
            for i in range(2, lws3 + 1):
                t1 = 'A' + str(i)
                td = 'D' + str(x)
                temp1 = gloVar.ws3[t1].value
                temp1 = temp1.upper()
                if temp1 == temp:
                    fy = 1
                    s = gloVar.ws3[t1].coordinate
                    g = 'G' + s[1:]
                    d = 'C' + s[1:]
                    e = 'D' + s[1:]
                    f = 'E' + s[1:]
                    u = 'B' + str(x)
                    for j, k in zip(cc, op):
                        ty = j + str(i)
                        tz = k + str(x)
                        ws1[tz] = gloVar.ws3[ty].value
                    catt = gloVar.ws3[g].value
                    try:
                        catt = catt.upper()
                    except:
                        pass
                    ws1[u] = catt
                    v = 'N' + str(x)
                    w = 'O' + str(x)
                    z = 'P' + str(x)
                    xc = gloVar.ws3[d].value
                    xd = gloVar.ws3[e].value
                    xe = gloVar.ws3[f].value
                    if xc is None:
                        xc = 0
                    if xd is None:
                        xd = 0
                    if xe is None:
                        xe = 0
                    if catt == 'LINK' or catt == 'MS':
                        if xc == 0 and xd == 0 and xe == 0:
                            gloVar.dictpwc[ws1[td].value] = temp
                            gloVar.sh[wy] = 0
                    ws1[v] = xc
                    ws1[w] = xd
                    ws1[z] = xe
                    break
        if fy == 0:
            td = 'D' + str(x)
            tempw = ws1[td].value
            tempw = tempw.upper()
            gloVar.cat[tempw] = temp
            gloVar.sh[wy] = 0


def breakupp(wy):
    alpha = ['A', 'B', 'D', 'E', 'F', 'G', 'H', 'I']
    beta = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    ws1 = gloVar.wb1[wy]
    #   Loop to break the primary list to board names and their corresponding quantity
    ws = gloVar.wb[wy]
    keys = {}
    dict = {}
    x = 12
    lws = ws.max_row + 1
    for col in range(12, lws):
        b = 'B' + str(x)
        cc = 'C' + str(x)
        if ws[b].value is None:
            break
        else:
            qt = ws[cc].value
            loco = ws[b].value
            dict[loco] = qt
            k = loco.split(' ')[0]
            keys[k] = qt
        x += 1
    x += 1
    z = x
    l = 2
    m = 0
    fl1 = 0
    for x in range(z, lws):
        t = 'B' + str(x)
        temp = ws[t].value
        if chk(temp) == 1:
            y = temp[13:]
            fb = y
            if y not in dict:
                print("BOOOOO")
                for w in dict.keys():
                    x = similar(y, w)
                    if x > 9:
                        fb = w
                        print(w)
                msg = messagebox.showinfo('BOM not Found', y + '  BOM not Found\n' + fb + '  Selected')
            ckc = 'BRD'
            qty = dict[fb]
            del (dict[fb])
            fl1 = 1
        elif chk(temp) == 2:
            y = temp[13:]
            fb = y
            if y not in dict:
                print("BOOOOO")
                for w in dict.keys():
                    x = similar(y, w)
                    if x > 0.9:
                        fb = w
                        print(w)
                msg = messagebox.showinfo('BOM not Found', y + '  BOM not Found\n' + fb + '  Selected')
            ckc = 'P' + y[:2]
            qty = dict[fb]
            del (dict[fb])
            fl1 = 1
        elif chk(temp) == 3:
            fl1 = 1
        else:
            for i in alpha:
                if i == 'B':
                    ac = 'C' + str(l)
                    ae = 'E' + str(l)
                    af = 'F' + str(l)
                    aj = 'J' + str(l)
                    ak = 'K' + str(l)
                    al = 'L' + str(l)
                    am = 'M' + str(l)
                    an = 'N' + str(l)
                    ao = 'O' + str(l)
                    ap = 'P' + str(l)
                    aq = 'Q' + str(l)
                    ar = 'R' + str(l)
                    s = 'S' + str(l)
                    at = 'T' + str(l)
                    ws1[ac] = ckc
                    ws1[aj] = qty
                    ws1[ak] = '=' + aj + '*' + ae
                    ws1[al] = '=' + af + '/3'
                    ws1[am] = '=' + ak + '*' + al
                    ws1[aq] = '=' + ak + '*' + an
                    ws1[ar] = '=' + ak + '*' + ao
                    ws1[s] = '=' + ak + '*' + ap
                    ws1[at] = '=' + an + '+' + ao + '+' + ap
                else:
                    tempa = beta[m] + str(x)
                    tempb = i + str(l)
                    ws1[tempb] = ws[tempa].value
                    m += 1
        if fl1 == 0:
            l += 1
            m = 0
        fl1 = 0
        x += 1
    x = 11
    fl3 = 0
    lws2 = gloVar.ws2.max_row + 1
    for col in range(11, lws2):
        t = 'B' + str(x)
        temp = gloVar.ws2[t].value
        if len(keys) == 0:
            break
        if temp is None:
            fl1 = 1
        elif temp.find("Feeder Code :") != -1:
            fl3 = 0
            for key in keys:
                if temp.find(key) != -1:
                    ckc = key
                    qty = keys[key]
                    fl3 = 1
            if fl3 == 1:
                del (keys[ckc])
                fl1 = 1
        elif fl3 == 1:
            for i in alpha:
                if i == 'B':
                    ac = 'C' + str(l)
                    ae = 'E' + str(l)
                    af = 'F' + str(l)
                    aj = 'J' + str(l)
                    ak = 'K' + str(l)
                    al = 'L' + str(l)
                    am = 'M' + str(l)
                    an = 'N' + str(l)
                    ao = 'O' + str(l)
                    ap = 'P' + str(l)
                    aq = 'Q' + str(l)
                    ar = 'R' + str(l)
                    s = 'S' + str(l)
                    at = 'T' + str(l)
                    ws1[ac] = ckc
                    ws1[aj] = qty
                    ws1[ak] = '=' + aj + '*' + ae
                    ws1[al] = '=' + af + '/3'
                    ws1[am] = '=' + ak + '*' + al
                    ws1[aq] = '=' + ak + '*' + an
                    ws1[ar] = '=' + ak + '*' + ao
                    ws1[s] = '=' + ak + '*' + ap
                    ws1[at] = '=' + an + '+' + ao + '+' + ap
                else:
                    tempa = beta[m] + str(x)
                    tempb = i + str(l)
                    ws1[tempb] = gloVar.ws2[tempa].value
                    m += 1
        if fl3 == 1:
            if fl1 == 0:
                l += 1
                m = 0
        fl1 = 0
        x += 1

    bup(wy)

    txtbox.insert(INSERT, wy + "  Successful!\n\t")
    print(wy + "  Successful!")
    if len(gloVar.cat) != 0:
        gloVar.sybr.append(wy)


def resize_image(event):
    new_width = event.width
    new_height = event.height
    image = copy_of_image.resize((new_width, new_height))
    photo = ImageTk.PhotoImage(image)
    label.config(image=photo)
    label.image = photo  # avoid garbage collection


def helloCallBack():
    global label1
    global file
    label1.destroy()
    gloVar.path = askopenfilename()
    print(gloVar.path.rsplit('/', 1)[-1])
    file = gloVar.path.rsplit('/', 1)[-1]
    label1 = Label(workzone, text=file, bg='#dadfe3', font=("Arial Rounded MT Bold", 8), foreground="#173d5c")
    label1.grid(row=1, column=2, sticky='ENWS', padx=(0, 5), pady=(5, 0))
    txtbox.delete(1.0, END)
    print(gloVar.path)
    if gloVar.path != '':
        txtbox.insert(INSERT, '\n  ' + file + ' selected.\n')
    else:
        txtbox.insert(INSERT, '\n  No File Selected.\n')


def breakup():
    if gloVar.path == 'No file selected\t':
        print('Please select a file.')
        msg = messagebox.showinfo('Please select a file.', gloVar.path)
    else:
        gloVar.dec(gloVar)
        print(gloVar.path.rsplit('/', 1)[-1])
        txtbox.insert(INSERT, '\n\n  Please wait loading....\n\n\t')
        i = 0
        print("Loading...")
        for s in gloVar.sheets:
            breakupp(s)
            run(i + 1)
            i += 1
    if len(gloVar.cat) != 0 or len(gloVar.dictpwc) != 0:
        main_menu.newWindowOpen()
    else:
        buttonbu = Button(root, image=conso, bg='#dadfe3', border='0', command=consolidate)
        buttonbu.grid(row=3, column=0, pady=(20, 20), sticky='NS')
        txtbox.insert(INSERT, "\n\n\t\tSuccess waiting for consolidation.....!\n\t")
    saveop()
    txtbox.insert(INSERT, "\n\n\tAll process Complete!!\n\t")


def saveop():
    file = gloVar.path.rsplit('/', 1)[-1]
    try:
        gloVar.wb1.save(filename='Output_' + file)
    except:
        msg = messagebox.showinfo('Close ' + 'Output_' + file, 'Please close the file and press OK to continue.')
        saveop()


def upd():
    x = e1.get()
    gloVar.pgoh = float(x)
    lb4 = Label(pgohframe, text=gloVar.pgoh, border='0', font=("Arial Rounded MT Bold", 11))
    lb4.grid(row=1, column=0, sticky='ENWS')
    buttonpg = Button(pgohframe, image=editt, border='0', command=editpgoh)
    buttonpg.grid(row=0, column=2, rowspan=2, padx=(5, 0))


def editpgoh():
    global e1
    e1 = Entry(pgohframe, width=10)
    e1.grid(row=1, column=0)
    b1 = Button(pgohframe, image=editt, border='0', command=upd)
    b1.grid(row=0, column=2, rowspan=2, padx=(5, 0))


root = Tk()
root.state('zoomed')
root.configure(background="#dadfe3")
root.title("TP Break Up")

# ========================================== Image URL =================================================================

title = resource_path("img/titletp.png")
img2 = resource_path("img/side.jpg")
lin = resource_path("img/line.png")
chfi = resource_path("img/choosefile.png")
ee1 = resource_path("img/edit.png")
chek = resource_path("img/check.png")
cosn = resource_path("img/CONS.png")

# =========================================== Opening Image ============================================================

image1 = Image.open(title)
imgr = Image.open(img2)
lini = Image.open(lin)
chck = Image.open(chek)
chofi = Image.open(chfi)
edi = Image.open(ee1)
cons = Image.open(cosn)

# =========================================== Resizing =================================================================

photo1 = ImageTk.PhotoImage(image1)
image1 = image1.resize((1400, 179))
copy_of_image1 = image1.copy()
photo1 = ImageTk.PhotoImage(image1)

imgri = ImageTk.PhotoImage(imgr)
imgr = imgr.resize((512, 384))
copy_of_imager = imgr.copy()
imgri = ImageTk.PhotoImage(imgr)

line = ImageTk.PhotoImage(lini)
lini = lini.resize((600, 15))
copy_of_imagel = lini.copy()
line = ImageTk.PhotoImage(lini)

checkk = ImageTk.PhotoImage(chck)
chck = chck.resize((166, 41))
copy_of_imagec = chck.copy()
checkk = ImageTk.PhotoImage(chck)

chosfi = ImageTk.PhotoImage(chofi)
chofi = chofi.resize((139, 37))
copy_of_imagech = chofi.copy()
chosfi = ImageTk.PhotoImage(chofi)

editt = ImageTk.PhotoImage(edi)
edi = edi.resize((40, 40))
copy_of_imagee = edi.copy()
editt = ImageTk.PhotoImage(edi)

conso = ImageTk.PhotoImage(cons)
cons = cons.resize((166, 41))
copy_of_imageco = cons.copy()
conso = ImageTk.PhotoImage(cons)

# =============================================== Frames ===============================================================

workzone = Frame(root, border='5', bg='grey')
workzone.grid(row=1, column=0, sticky='ENW', padx=(80, 10), pady=(30, 0), rowspan=2)

pgohframe = Frame(root, border='5', bg='grey')
pgohframe.grid(row=1, column=1, padx=(10, 80), pady=(30, 0))

progress = Frame(root, border='5', bg='#dadfe3')
progress.grid(row=5, column=0, sticky='NS', padx=(10, 80))

group1 = LabelFrame(root, text="Details", padx=5, pady=5, bg='#dadfe3')
group1.grid(row=7, column=0, padx=80, pady=(50, 45), sticky=E + W + N + S)

# ============================================ Logo and Side Image =====================================================

title = Label(root, bg='#dadfe3', image=photo1)
title.grid(row=0, column=0, columnspan=2, sticky=N, pady=30)

sideimage = Label(root, bg='#dadfe3', image=imgri)
sideimage.grid(row=2, column=1, sticky='ENWS', padx=(0, 100), pady=(0, 20), rowspan=7)

# ============================================ Buttons and Labels ======================================================

lb3 = Label(pgohframe, text='  PGOH Value  ', bg='#a0a0a0', border='0', font=("Arial Rounded MT Bold", 11),
            foreground="white")
lb3.grid(row=0, column=0, sticky='ENWS', pady=(0, 5))
lb4 = Label(pgohframe, text=gloVar.pgoh, border='0', font=("Arial Rounded MT Bold", 11))
lb4.grid(row=1, column=0, sticky='ENWS')
buttonpg = Button(pgohframe, image=editt, border='0', command=editpgoh)
buttonpg.grid(row=0, column=2, rowspan=2, padx=(5, 0))

lb1 = Label(workzone, text='Select Files', bg='#a0a0a0', font=("Arial Rounded MT Bold", 15), foreground="white")
lb1.grid(row=0, column=0, sticky='ENWS', padx=(0, 5))
lb2 = Label(workzone, text='Selected File', bg='#a0a0a0', font=("Arial Rounded MT Bold", 15), foreground="white")
lb2.grid(row=0, column=1, sticky='ENWS', padx=(0, 5), columnspan=2)
lb21 = Label(workzone, text='Status', bg='#a0a0a0', font=("Arial Rounded MT Bold", 15), foreground="white")
lb21.grid(row=0, column=3, sticky='ENWS')

lb11 = Label(workzone, text='TP Detailed Report', bg='#dadfe3', border='5', font=("Arial Rounded MT Bold", 12),
             foreground="#173d5c")
lb11.grid(row=1, column=0, sticky='ENWS', padx=(0, 5), pady=(5, 0))
label1 = Label(workzone, text=gloVar.path, bg='#dadfe3', font=("Arial Rounded MT Bold", 8), foreground="#173d5c")
label1.grid(row=1, column=2, sticky='ENWS', padx=(0, 5), pady=(5, 0))
buttoncf = Button(workzone, image=chosfi, bg='#dadfe3', border='0', command=helloCallBack)
buttoncf.grid(row=1, column=1, padx=(0, 5), pady=(5, 0), sticky='EW')
lb1 = Label(workzone, text='', bg='#dadfe3', font=("Arial Rounded MT Bold", 12), foreground="green")
lb1.grid(row=1, column=3, sticky='ENWS', pady=(5, 0))

buttonbu = Button(root, image=checkk, bg='#dadfe3', border='0', command=breakup)
buttonbu.grid(row=3, column=0, pady=(20, 20), sticky='NS')

lbl = Label(root, image=line, bg='#dadfe3')
lbl.grid(row=4, column=0, sticky='ENWS', pady=0)

lbp = Label(progress, text='Progress: ', bg='#dadfe3')
lbp.grid(row=0, column=0, sticky='ENWS', pady=0)
progressBar = ttk.Progressbar(progress, orient="horizontal", length=286, mode="determinate")
progressBar.grid(row=0, column=1, pady=5)

lbl2 = Label(root, image=line, bg='#dadfe3')
lbl2.grid(row=6, column=0, sticky='ENWS', pady=0)

# Create the textbox
txtbox = scrolledtext.ScrolledText(group1, width=10, height=10)
txtbox.grid(row=0, column=0, padx=0, pady=(0, 0), sticky='NSEW')

# ============================================= Column Weights =========================================================

pgohframe.columnconfigure(0, weight=1)
pgohframe.columnconfigure(1, weight=1)
root.columnconfigure(0, weight=15)
root.columnconfigure(1, weight=2)
workzone.columnconfigure(0, weight=2)
workzone.columnconfigure(1, weight=1)
workzone.columnconfigure(2, weight=3)
workzone.columnconfigure(3, weight=1)
group1.rowconfigure(0, weight=1)
group1.columnconfigure(0, weight=1)
root.rowconfigure(7, weight=1)

# ======================================================================================================================
main_menu = main()
root.mainloop()
